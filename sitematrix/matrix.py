from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()
#setting up output file
dest_filename = 'output.xlsx'
ws1 = wb.active
ws1.title = "doctors"
ws2 = wb.create_sheet(title="conditions")
ws3 = wb.create_sheet(title="treatment")
ws4 = wb.create_sheet(title="test")
ws5 = wb.create_sheet(title="clinical-trials")

ws1['A1']="Doctor Path"
ws2['A1']="Conditions"
ws2['B1']="Conditions Path"
ws3['A1']="Treatment Path"
ws4['A1']="Test Path"
ws5['A1']="Trial Path"
#get input from user
clinic = raw_input("Please enter clinic path: (ex: /content/shc/en/medical-clinics/injury-prevention-program/community-partners-resources.html")

#finding initial info from clinic report
readbook = load_workbook('clinics-report-exclemation.xlsx')
cm=readbook.active
cm_size=cm.max_row
i=1
for col in cm.iter_cols(min_row=2, min_col=2, max_col=2, max_row=cm_size):
     for cell in col:
         i=i+1
         if (clinic==cell.value):
             dcell="G"+str(i)
             doctors=cm[dcell].value
             concell = "I" + str(i)
             conditions = cm[concell].value
             conpathcell = "J" + str(i)
             conditionpaths = cm[conpathcell].value
             #get other values here
             break
#break out doctorpath list
try:
    doctor_list=doctors.split('!')
    i=2
    for doctor in doctor_list:
        current="A"+str(i)
        #print(doctor)
        ws1[current]=doctor
        i=i+1
except:
    ws1['A2']="No Doctors"
#break out conditions
try:
    condition_list=conditions.split('!')
    i=2
    for condition in condition_list:
        current="A"+str(i)
        #print(doctor)
        ws2[current]=condition
        i=i+1
except:
    ws2['A2'] = "No Conditions"
#break out condition paths
try:
    conditionpath_list=conditionpaths.split('!')
    i=2
    for conditionpath in conditionpath_list:
        current="B"+str(i)
        #print(doctor)
        ws2[current]=conditionpath
        i=i+1
except:
    ws2['B2'] = "No Condition Paths"



wb.save(filename=dest_filename)